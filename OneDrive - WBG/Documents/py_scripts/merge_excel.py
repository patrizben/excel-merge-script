import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# ── CONFIGURATION ──────────────────────────────────────────────────────────────
INPUT_FOLDER  = r"C:\Users\wb601985\Downloads\Charts"  # <-- Change this
OUTPUT_FILE = r"C:\Users\wb601985\Downloads\Charts\PEUJun2026_storylinecharts.xlsx"  # <-- Change this
# ───────────────────────────────────────────────────────────────────────────────


merged_wb = openpyxl.Workbook()
merged_wb.remove(merged_wb.active)  # Remove the default blank sheet

summary_data = []  # Will hold (tab_name, source_filename) tuples

# ── STEP 1: Loop through all Excel files in the folder ─────────────────────────
import re

def natural_sort_key(filename):
    return [int(c) if c.isdigit() else c.lower() for c in re.split(r'(\d+)', filename)]

for filename in sorted(os.listdir(INPUT_FOLDER), key=natural_sort_key):
    if not filename.endswith((".xlsx", ".xlsm")):
        continue

    source_path = os.path.join(INPUT_FOLDER, filename)
    source_wb   = openpyxl.load_workbook(source_path, data_only=True)
    base_name   = os.path.splitext(filename)[0]  # Filename without extension

    for sheet_name in source_wb.sheetnames:
        source_ws = source_wb[sheet_name]

        # Build tab name: "Filename - SheetName", trimmed to 31 chars (Excel limit)
        raw_tab_name = f"{base_name} - {sheet_name}"
        tab_name     = raw_tab_name[:31]

        # Handle duplicate tab names
        existing_titles = [ws.title for ws in merged_wb.worksheets]
        counter = 1
        unique_tab = tab_name
        while unique_tab in existing_titles:
            suffix   = f" ({counter})"
            unique_tab = tab_name[:31 - len(suffix)] + suffix
            counter += 1

        new_ws = merged_wb.create_sheet(title=unique_tab)

        # Copy all cell values and basic formatting
        for row in source_ws.iter_rows():
            for cell in row:
                new_cell       = new_ws.cell(row=cell.row, column=cell.column)
                new_cell.value = cell.value

        # Copy column widths
        for col_letter, col_dim in source_ws.column_dimensions.items():
            new_ws.column_dimensions[col_letter].width = col_dim.width

        # Add link to Summary in A1
        if new_ws['A1'].value is not None and new_ws['A1'].value != "":
            new_ws.insert_rows(1)
        new_ws['A1'].value = "Back to Summary"
        new_ws['A1'].hyperlink = "#'Summary'!A1"
        new_ws['A1'].font = Font(color="0000FF", underline="single")

        summary_data.append((unique_tab, filename))
        print(f"  Added: '{unique_tab}' from '{filename}'")

# ── STEP 2: Create the Summary tab at the front ────────────────────────────────
summary_ws = merged_wb.create_sheet(title="Summary", index=0)

summary_ws["A1"] = "Summary of Merged Sheets"
summary_ws["A1"].font = Font(bold=True, size=14)
summary_ws["A2"] = ""

summary_ws["A3"] = "Sheet Tab"
summary_ws["B3"] = "Source File"
summary_ws["A3"].font = Font(bold=True)
summary_ws["B3"].font = Font(bold=True)

for i, (tab_name, source_file) in enumerate(summary_data, start=4):
    # Clickable hyperlink to the tab
    cell = summary_ws.cell(row=i, column=1)
    cell.value     = tab_name
    cell.hyperlink = f"#'{tab_name}'!A1"
    cell.font      = Font(color="0000FF", underline="single")

    summary_ws.cell(row=i, column=2).value = source_file

summary_ws.column_dimensions["A"].width = 40
summary_ws.column_dimensions["B"].width = 35

# ── STEP 3: Save the merged workbook ───────────────────────────────────────────
os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
merged_wb.save(OUTPUT_FILE)
print(f"\nDone! Merged workbook saved to: {OUTPUT_FILE}")
